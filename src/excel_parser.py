"""
Excel 解析模块 v2

新逻辑：
  1. 从 ai数量需求汇总 副表读取每组每产品的需求数量
  2. 从各组副表按产品段落扫描模版行
  3. 循环模版补齐到需求数量
  4. 每个产品独立输出一个 JSON 文件（{组名}_{产品名}.json）
"""
import json
import re
import datetime
import openpyxl
from pathlib import Path


def _clean_cell(cell) -> str:
    """清理单元格值：None→'/'，换行符→空格"""
    if cell is None:
        return '/'
    if isinstance(cell, datetime.time):
        return f'{cell.hour}:{cell.minute:02d}'
    # Excel/钉钉把 "9:16" 这种比例自动识别为时间，存成 0~1 浮点小数
    # （如 0.386111111... = 9*60+16 分钟 / 1440 → 还原为 "9:16"）
    # 启发式：repr 长度 > 5 排除人为输入的 0.5 / 0.25 等短小数
    if isinstance(cell, float) and 0 < cell < 1 and len(repr(cell)) > 5:
        total_min_f = cell * 1440
        total_min = round(total_min_f)
        if abs(total_min_f - total_min) < 0.05:
            h, m = divmod(total_min, 60)
            return f'{h}:{m}'
    val = str(cell).strip()
    if not val:
        return '/'
    return val.replace('\n', ' ').replace('\r', '')


def _is_header_row(row) -> bool:
    """是否为列标题行（首列含 '日期' 或 'hu' 关键词，或含 '人物a' 等典型列名）"""
    _DATE_KEYWORDS = {'日期', 'hu'}
    _COL_KEYWORDS = {'人物a', '场景', '道具', '输入指令台词'}
    first = str(row[0]).strip().lower() if row[0] is not None else ''
    if first in _DATE_KEYWORDS:
        return True
    return any(str(c).strip() in _COL_KEYWORDS for c in row if c is not None)


def _is_title_row(row) -> bool:
    """是否为产品标题行：首列非空字符串、非日期类型、其余列基本为空"""
    first = row[0]
    if first is None:
        return False
    if isinstance(first, (float, int, datetime.date, datetime.datetime, datetime.time)):
        return False
    s = str(first).strip()
    if len(s) < 2 or s == '日期':
        return False
    if _is_header_row(row):
        return False
    non_empty = sum(1 for c in row[1:] if c is not None and str(c).strip())
    return non_empty <= 2


def _row_to_str(row: tuple, col_count: int) -> str:
    """将数据行转成 ' // ' 拼接字符串，去掉末尾空列"""
    values = [_clean_cell(row[i] if i < len(row) else None) for i in range(col_count)]
    while values and values[-1] == '/':
        values.pop()
    return ' // '.join(values)


def _scan_sections(rows: list) -> list:
    """
    扫描工作表，提取所有产品段落。
    返回: [{'title': str, 'header_str': str, 'col_count': int, 'templates': [str]}]
    """
    sections = []
    cur_title = None
    cur_header_str = None
    cur_col_count = 0
    cur_templates = []

    for row in rows:
        if _is_title_row(row):
            if cur_title and cur_header_str:
                sections.append({
                    'title': cur_title,
                    'header_str': cur_header_str,
                    'col_count': cur_col_count,
                    'templates': cur_templates,
                })
            cur_title = str(row[0]).strip()
            cur_header_str = None
            cur_col_count = 0
            cur_templates = []
        elif _is_header_row(row):
            headers = [str(c).strip() for c in row if c is not None and str(c).strip()]
            cur_header_str = ' // '.join(headers)
            cur_col_count = len(headers)
        elif cur_header_str:
            # 过滤"空壳行"：甲方常留下只有日期/比例而其他列都是 / 或空的占位行
            # 必须至少有 2 列实质内容（排除日期列、排除 / 占位），才算有效模版
            meaningful = sum(
                1 for c in row[1:]
                if c is not None and str(c).strip() not in ('', '/')
            )
            if meaningful >= 2:
                s = _row_to_str(row, cur_col_count)
                if s:
                    cur_templates.append(s)

    # 最后一个段落
    if cur_title and cur_header_str:
        sections.append({
            'title': cur_title,
            'header_str': cur_header_str,
            'col_count': cur_col_count,
            'templates': cur_templates,
        })
    return sections


def _norm_group(name: str) -> str:
    """
    组名归一化：去空白 + 去尾部「组」字 + 再去空白。
    用于容错匹配（汇总表写「余相波」也能命中工作簿 sheet「余相波组」）。
    例：'余相波组' / '余相波' / ' 余相波 组 ' → 全部归一为 '余相波'
    """
    s = str(name or '').strip()
    while s.endswith('组'):
        s = s[:-1].strip()
    return s


def _parse_demand(ws) -> tuple:
    """
    解析 ai数量需求汇总 副表。
    返回: (date_str, {'何欢组': [('redhalo生物素', 40), ...], ...})
    date_str 取汇总表第一个非空日期值（如 '5.16'）

    列位置通过表头文字动态识别（兼容甲方调整列顺序/插入辅助列）：
      日期    ← 表头含「日期」
      组      ← 表头精确等于「组」或含「分组」/「组名」
      产品    ← 表头含「品类」/「产品」/「品名」
      数量    ← 表头含「需求数量」/「数量」
    """
    demand = {}
    current_group = None
    batch_date    = None
    idx_date = idx_group = idx_prod = idx_qty = None

    def _match_header(cells: list) -> bool:
        nonlocal idx_date, idx_group, idx_prod, idx_qty
        date_i = group_i = prod_i = qty_i = None
        for i, c in enumerate(cells):
            if c is None:
                continue
            s = str(c).strip()
            if not s:
                continue
            if date_i is None and ('日期' in s):
                date_i = i
            if group_i is None and (s == '组' or '分组' in s or '组名' in s):
                group_i = i
            if prod_i is None and ('品类' in s or '产品' in s or '品名' in s):
                prod_i = i
            if qty_i is None and ('需求数量' in s or s == '数量' or s.endswith('数量')):
                qty_i = i
        # 至少要识别到「组」和「数量」两列才算表头
        if group_i is not None and qty_i is not None:
            idx_date, idx_group, idx_prod, idx_qty = date_i, group_i, prod_i, qty_i
            return True
        return False

    header_found = False
    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        if not header_found:
            if _match_header(cells):
                header_found = True
            continue
        date_val = cells[idx_date] if (idx_date is not None and idx_date < len(cells)) else None
        group    = cells[idx_group] if (idx_group is not None and idx_group < len(cells)) else None
        product  = cells[idx_prod]  if (idx_prod  is not None and idx_prod  < len(cells)) else None
        quantity = cells[idx_qty]   if (idx_qty   is not None and idx_qty   < len(cells)) else None

        if date_val and str(date_val).strip():
            if batch_date is None:
                batch_date = str(date_val).strip()
        if group and str(group).strip():
            current_group = str(group).strip()
        # 数量可能是 int / float / "40" 字符串
        qty_int = None
        if isinstance(quantity, (int, float)):
            qty_int = int(quantity)
        elif isinstance(quantity, str) and quantity.strip().isdigit():
            qty_int = int(quantity.strip())
        if current_group and product and qty_int and qty_int > 0:
            demand.setdefault(current_group, []).append((str(product).strip(), qty_int))

    # 容错匹配：额外构建一份「归一化组名 → 原始组名」映射，便于按 sheet 反查
    norm_index = {}
    for raw_key in demand.keys():
        nk = _norm_group(raw_key)
        norm_index.setdefault(nk, raw_key)

    return (batch_date or '未知日期', demand, norm_index)


def _normalize(s: str) -> str:
    """规范化产品名用于模糊匹配：去括号内容、小写、去空格"""
    s = re.sub(r'[（(][^）)]*[）)]', '', s)
    return s.lower().replace(' ', '').strip()


def _match_section(product_name: str, sections: list) -> dict:
    """
    模糊匹配需求产品名 → 模版段落。
    规则：
      1) 先取「归一化后完全相等」的所有候选，存在则返回最后一个（=Excel 中位置最靠后 = 最新版本）
      2) 否则退回「子串包含」匹配，同样取最后一个
      3) 都没有则兜底返回第一个段落
    背景：甲方常把同一产品的"需要加字幕/不需要加字幕"两份段落都留在 sheet 里，
         旧的在上、新的在下。原"从前往后第一个命中"会拿到历史段落，导致清洗
         使用了过时模板。
    """
    norm = _normalize(product_name)
    exact = [s for s in sections if _normalize(s['title']) == norm]
    if exact:
        if len(exact) > 1:
            cands = ' | '.join(s['title'] for s in exact)
            print(f'    ⚠️  「{product_name}」匹配到 {len(exact)} 个同名段落，取最后一个；候选：{cands}')
        return exact[-1]
    loose = [s for s in sections
             if (norm in _normalize(s['title']) or _normalize(s['title']) in norm)]
    if loose:
        if len(loose) > 1:
            cands = ' | '.join(s['title'] for s in loose)
            print(f'    ⚠️  「{product_name}」子串匹配到 {len(loose)} 个段落，取最后一个；候选：{cands}')
        return loose[-1]
    return sections[0] if sections else None


# ====================================================================
# 颜色洗牌相关常量与工具
# 长词在前，避免 "深红" 被切成 "红"
# ====================================================================
_COLOR_WORDS = [
    # 复合色（必须放最前）
    "卡其", "姜黄", "墨绿", "藏青", "藏蓝", "玫红", "酒红", "宝蓝", "天蓝",
    "湖蓝", "海蓝", "米白", "米色", "杏色", "奶白", "象牙", "驼色",
    "深红", "深蓝", "深灰", "深绿", "深棕", "深紫",
    "浅蓝", "浅灰", "浅黄", "浅绿", "浅粉", "浅紫", "淡黄", "淡紫",
    "青绿", "青蓝", "青灰", "青色",
    "灰蓝", "灰绿",
    # 基本色
    "黑", "白", "灰", "蓝", "红", "绿", "黄", "橙", "紫",
    "粉", "棕", "金", "银", "青", "褐", "栗",
]
# 颜色 + 可选"色" + 后面跟服装/配饰词，才视为有效"服装颜色"
# 排除"颜色 + 物品"中的非服装情况（如"白纸""黑笔"）
_CLOTH_WORDS = (
    "衬衫|上衣|T恤|短袖|长袖|外套|马甲|大衣|风衣|羽绒服|夹克|"
    "卫衣|毛衣|裙子|连衣裙|半身裙|裙|裤子|长裤|短裤|牛仔裤|"
    "西装|套装|工装|制服|工作服|护士服|中医服|中式上衣|"
    "睡衣|睡裙|睡袍|家居服|"
    "大褂|长袍|长衫|罩衫|中山装|唐装|旗袍|汉服|"
    "polo衫|Polo衫|POLO衫|polo领|Polo领|POLO领|polo|Polo|POLO|"
    "围裙|围巾|领带|帽子|手套|鞋|"
    "眼镜|墨镜|框架|框"
)
# 黑名单：这些"颜色 + 词"是固定搭配，不允许换颜色（白大褂=约定俗成的医生制服）
_COLOR_BLACKLIST = {"白大褂"}

# 编译用于扫描"颜色+服装"位置的正则
# 支持两种形态：
#   1) 紧贴：'黑色衬衫' / '蓝衬衫'   → 颜色 + 可选"色" + 服装词
#   2) 隔修饰词：'蓝色polo领衬衫' / '白色医生大褂' → 颜色 + "色" + 1~8 个非标点修饰字 + 服装词
# 关键约束：形态 2 必须有"色"字才允许跨字符，避免 '白纸黑色衬衫' 误把"白"当成衬衫的颜色
import re as _re
_COLOR_GROUP = "(?:" + "|".join(_COLOR_WORDS) + ")"
_CLOTH_PATTERN = _re.compile(
    rf"({_COLOR_GROUP})(?:色[^，,。.！!？?；;、\s]{{1,10}}|色?)(?={_CLOTH_WORDS})"
)
# 纯颜色扫描（用于构建颜色池）：与替换 pattern 完全一致，确保颜色池覆盖率
_COLOR_SCAN_PATTERN = _CLOTH_PATTERN


def _extract_color_pool(texts: list) -> list:
    """
    从一批人物描述文本中扫描所有出现过的"服装颜色"，去重保序返回。
    只收集出现在 颜色+服装 模式里的颜色（避免把"白纸"的"白"误纳入）。
    """
    pool, seen = [], set()
    for t in texts:
        if not t:
            continue
        for m in _COLOR_SCAN_PATTERN.finditer(t):
            # 黑名单：仅当颜色起始位置正好就是黑名单短语（如 '白大褂'）时不入池
            # 不用滑窗 substring 判定，避免 '白色短袖白大褂' 这种共存场景误伤
            start = m.start(1)
            if any(t.startswith(bl, start) for bl in _COLOR_BLACKLIST):
                continue
            c = m.group(1)
            if c not in seen:
                seen.add(c)
                pool.append(c)
    return pool


def _swap_colors_in_text(text: str, pool: list, rng) -> str:
    """
    把文本里所有"颜色+服装"位置的颜色随机替换为颜色池里另一个（≠原色，避免替换无效）。
    黑名单（如"白大褂"）跳过不换。从右往左替换，避免位置偏移。
    """
    if not text or not pool or len(pool) < 2:
        return text
    matches = list(_CLOTH_PATTERN.finditer(text))
    if not matches:
        return text
    new_text = text
    for m in reversed(matches):
        start, end = m.start(1), m.end(1)
        old_color = m.group(1)
        # 黑名单：仅当颜色 *起始位置* 正好等于黑名单短语开头时跳过（如 '白大褂'）
        # 不用滑窗 substring 判定，避免 '白色短袖白大褂' 这种共存场景把"短袖"也保护了
        if any(new_text.startswith(bl, start) for bl in _COLOR_BLACKLIST):
            continue
        # 候选池：去掉原色后随机选
        candidates = [c for c in pool if c != old_color]
        if not candidates:
            continue
        new_color = rng.choice(candidates)
        new_text = new_text[:start] + new_color + new_text[end:]
    return new_text


# ====================================================================
# 列分组识别
# ====================================================================
def _classify_columns(headers: list) -> tuple:
    """
    根据表头把列分成三类，返回 (text_group, visual_group, char_group)：
      - text_group   : 台词/字幕  → 整组锁定到同一原始行抽
      - visual_group : 人物a/b、场景、道具、动作、镜头、时长   → 整组锁定到同一原始行抽
      - char_group   : 人物a/b 的列索引子集（用于颜色洗牌目标）
    其余列保持独立随机（如"日期"）。
    备注：时长归视觉组，因为"同一镜头 → 同一时长"在语义上是配套的。
    """
    text_group, visual_group, char_group = [], [], []
    for i, h in enumerate(headers):
        if ("台词" in h) or ("字幕" in h):
            text_group.append(i)
        elif ("人物" in h) or ("角色" in h):
            visual_group.append(i)
            char_group.append(i)
        elif any(k in h for k in ("场景", "道具", "动作", "镜头", "参考图", "环境", "背景", "时长")):
            visual_group.append(i)
    return text_group, visual_group, char_group



def count_unique_text_rows(templates: list, header_str: str) -> int:
    """统计模版行中唯一台词组合的数量（即"有几套不重复的台词"）。

    只按"台词"列去重（排除"字幕"列），因为字幕是台词的校正版，
    不同字幕但相同台词视为同一套。
    - 有台词列（含"台词"且不含"字幕"的列）：按这些列组合去重计数
    - 无台词列：返回去重后的总行数（整行视为台词）
    """
    if not templates:
        return 0
    headers = [h.strip() for h in header_str.split(' // ')]
    # 只取含"台词"且不含"字幕"的列（排除"最终字幕的台词"这类）
    lyric_idx = [i for i, h in enumerate(headers) if '台词' in h and '字幕' not in h]
    if not lyric_idx:
        # 兜底：用全部 text_group（含字幕）
        text_idx, _, _ = _classify_columns(headers)
        lyric_idx = text_idx
    if not lyric_idx:
        return len(set(templates))
    seen = set()
    for row_str in templates:
        parts = row_str.split(' // ')
        parts += ['/'] * (len(headers) - len(parts))
        t = tuple(parts[i] if i < len(parts) else '/' for i in lyric_idx)
        seen.add(t)
    return len(seen)


# 时长列兜底常量
_DEFAULT_DURATION = '10'
_DURATION_RE = _re.compile(r'^\d+(?:\.\d+)?$')


def _fix_duration_value(v: str) -> str:
    """时长单元格兜底：空 / 占位 '/' / 非数字 → '10'。已为合法数字则原样返回。"""
    s = (v or '').strip()
    if not s or s == '/' or not _DURATION_RE.match(s):
        return _DEFAULT_DURATION
    return s


def _random_combine(templates: list, header_str: str, count: int, seed: int = None) -> list:
    """
    从模版组合生成 count 条新数据，新算法：
      1) 视觉组（人物a/b + 场景 + 道具 + 动作 + 镜头参考图）整组锁定到同一原始行抽
         → 保证场景与人物搭配合理（不会出现"睡衣+医生办公室"）
      2) 台词组（台词/字幕）整组锁定到同一原始行抽
         → 保证文案完整不串行
      3) 颜色洗牌：在人物 a/b 列上随机替换衣服颜色（从原模版颜色池中选）
         → 在合理搭配基础上额外补充多样性
         → 黑名单（如"白大褂"）保护，不会乱换
      其余未分类列（如"日期"）各自独立随机。
    """
    import random
    if not templates or count == 0:
        return []

    rng = random.Random(seed)  # seed=None 完全随机；指定则可复现
    headers = [h.strip() for h in header_str.split(' // ')]
    col_count = len(headers)

    text_idx, visual_idx, char_idx = _classify_columns(headers)
    text_set, visual_set = set(text_idx), set(visual_idx)
    char_set = set(char_idx)
    # 时长列索引（用于空值兜底为 10；可能 0 / 1 / 多列）
    duration_idx = [i for i, h in enumerate(headers) if '时长' in h]

    # 按列拆出所有模版值，同时收集整组（保持同行关联）
    columns      = [[] for _ in range(col_count)]
    text_rows    = []  # 台词组：[(列x值, 列y值), ...]
    visual_rows  = []  # 视觉组：[(列a值, 列b值, ...), ...]
    char_texts   = []  # 人物列原文集合（用于建颜色池）
    for row_str in templates:
        parts = row_str.split(' // ')
        parts += ['/'] * (col_count - len(parts))  # 补齐
        for i, v in enumerate(parts[:col_count]):
            columns[i].append(v)
        if text_idx:
            text_rows.append(tuple(parts[i] for i in text_idx))
        if visual_idx:
            visual_rows.append(tuple(parts[i] for i in visual_idx))
        for i in char_idx:
            char_texts.append(parts[i])

    # 构建颜色池：仅来自原模版人物列（方案 A — 安全的）
    color_pool = _extract_color_pool(char_texts)

    # 台词队列：只按"台词"列（不含"字幕"列）去重，字幕不同但台词相同视为同一套
    # → 保证生成的前 K 条（K = 唯一台词数）各有不同台词，超出后再随机
    # lyric_idx_local：text_idx 中仅含"台词"且不含"字幕"的列（用于去重 key）
    lyric_idx_local = [i for i in text_idx if '台词' in headers[i] and '字幕' not in headers[i]]
    if not lyric_idx_local:
        lyric_idx_local = text_idx  # 兜底：没有纯台词列时仍用全部文本列

    text_queue: list = []
    if text_rows:
        # 以台词列为 key 去重，value 保留完整 text_row（含字幕），取首次出现
        seen_lyric: dict = {}
        for tr in text_rows:
            # tr 的顺序与 text_idx 对应
            key = tuple(tr[text_idx.index(i)] for i in lyric_idx_local if i in text_idx)
            if key not in seen_lyric:
                seen_lyric[key] = tr
        unique_texts = list(seen_lyric.values())
        shuffled = unique_texts[:]
        rng.shuffle(shuffled)
        text_queue = shuffled[:]

    # 生成 count 条
    result = []
    for _ in range(count):
        new_parts = ['/'] * col_count
        # 视觉组整组绑定
        if visual_rows:
            v = rng.choice(visual_rows)
            for i, val in zip(visual_idx, v):
                new_parts[i] = val
        # 台词组：优先从队列取（保证不重复），队列耗尽后随机
        if text_rows:
            t = text_queue.pop(0) if text_queue else rng.choice(text_rows)
            for i, val in zip(text_idx, t):
                new_parts[i] = val
        # 其余未分类列：独立随机
        for i in range(col_count):
            if i in visual_set or i in text_set:
                continue
            new_parts[i] = rng.choice(columns[i]) if columns[i] else '/'
        # 颜色洗牌：仅对人物 a/b 列做颜色替换
        for i in char_set:
            new_parts[i] = _swap_colors_in_text(new_parts[i], color_pool, rng)
        # 时长兜底：空 / '/' / 非数字 → '10'
        for i in duration_idx:
            new_parts[i] = _fix_duration_value(new_parts[i])
        result.append(' // '.join(new_parts))

    return result


def parse_excel(excel_path: str, output_dir: str = 'data/input/parsed') -> dict:
    """
    主入口：解析 Excel，按需求数量扩展模版，每产品输出一个 JSON 文件。
    文件夹命名：{output_dir}/{日期}/
    文件命名：{日期}_{副表名}_{产品名}.json
    返回: {file_key: json_data}
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # 1. 读取需求汇总（含日期 + 归一化索引）
    batch_date, demand, demand_norm_index = _parse_demand(wb['ai数量需求汇总'])
    total_needed = sum(q for prods in demand.values() for _, q in prods)
    print(f'\n📅 批次日期：{batch_date}')
    print(f'📋 需求汇总：{len(demand)} 个组，合计 {total_needed} 条视频')
    for grp, prods in demand.items():
        for prod, qty in prods:
            print(f'   {grp} | {prod} | {qty}条')

    # 输出目录：data/input/parsed/{日期}/
    out_dir = Path(output_dir) / batch_date
    out_dir.mkdir(parents=True, exist_ok=True)
    print(f'📁 输出目录：{out_dir}')

    # 2. 逐组解析模版并扩展
    all_results = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == 'ai数量需求汇总':
            continue
        # 容错匹配：先按原名命中；不中则按归一化（去尾「组」字）兜底
        if sheet_name in demand:
            demand_key = sheet_name
        else:
            nk = _norm_group(sheet_name)
            demand_key = demand_norm_index.get(nk)
            if demand_key is None:
                print(f'\n  ⏭️  [{sheet_name}] 需求表无对应，跳过')
                continue
            print(f'\n  🔁 [{sheet_name}] 需求表写作「{demand_key}」，按归一化「{nk}」匹配成功')

        rows = list(wb[sheet_name].iter_rows(values_only=True))
        sections = _scan_sections(rows)
        if not sections:
            print(f'\n  ⚠️  [{sheet_name}] 未找到产品段落，跳过')
            continue

        print(f'\n📂 [{sheet_name}] 找到 {len(sections)} 个产品段落')
        for product_name, quantity in demand[demand_key]:
            sec = _match_section(product_name, sections)
            if not sec:
                print(f'  ⚠️  [{sheet_name}] {product_name} 无匹配段落，跳过')
                continue

            expanded = _random_combine(sec['templates'], sec['header_str'], quantity)
            json_data = [sec['title'], sec['header_str']] + expanded

            prod_clean = re.sub(r'[（(][^）)]*[）)]', '', product_name).strip()
            # 新命名：日期_副表名_产品名
            file_key = f'{batch_date}_{sheet_name}_{prod_clean}'
            out_file = out_dir / f'{file_key}.json'
            with open(out_file, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, ensure_ascii=False, indent=2)

            tmpl_count = len(sec['templates'])
            print(f'  ✅  {file_key}：{tmpl_count}条模版 → 扩展{quantity}条 → {out_file.name}')
            all_results[file_key] = json_data

    print(f'\n✅ 完成：共 {len(all_results)} 个 JSON 文件，保存在 {out_dir}\n')
    return all_results


# ====================================================================
# 手动模式：扫描产品（不读需求表、不扩展）
# ====================================================================
def scan_products(excel_path: str) -> list:
    """
    扫描 Excel 中所有非「ai数量需求汇总」的工作表，返回所有产品段落（不去重、不扩展）。
    供前端展示给用户手动指定数量。
    返回: [
      {
        'sheet': str,           # 工作表名（组名）
        'title': str,           # 产品段落原始标题（含括号备注）
        'product_name': str,    # 去括号的纯产品名
        'header_str': str,      # 表头字符串
        'col_count': int,
        'templates': [str],     # 原始模版行（' // ' 拼接）
        'template_count': int,
      }, ...
    ]
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    results = []
    for sheet_name in wb.sheetnames:
        if sheet_name == 'ai数量需求汇总':
            continue
        rows = list(wb[sheet_name].iter_rows(values_only=True))
        sections = _scan_sections(rows)
        if not sections:
            continue
        for sec in sections:
            if not sec.get('templates'):
                continue
            title = sec['title']
            prod_clean = re.sub(r'[（(][^）)]*[）)]', '', title).strip()
            results.append({
                'sheet': sheet_name,
                'title': title,
                'product_name': prod_clean,
                'header_str': sec['header_str'],
                'col_count': sec['col_count'],
                'templates': sec['templates'],
                'template_count': len(sec['templates']),
            })
    return results


def expand_product(templates: list, header_str: str, count: int,
                   batch_date: str, sheet_name: str, product_title: str,
                   output_dir: str = 'data/input/parsed',
                   seed: int = None) -> dict:
    """
    单产品扩展 + 写盘。
    给定原始模版 + 数量，调用 _random_combine 扩展，落地 JSON。
    返回: {'file_key': str, 'out_path': str, 'rows': [str], 'header_str': str, 'title': str}
    """
    prod_clean = re.sub(r'[（(][^）)]*[）)]', '', product_title).strip()
    # 文件 key 与现有 parse_excel 保持一致：日期_组名_产品名
    file_key = f'{batch_date}_{sheet_name}_{prod_clean}'
    expanded = _random_combine(templates, header_str, int(count), seed=seed)
    json_data = [product_title, header_str] + expanded

    out_dir = Path(output_dir) / batch_date
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / f'{file_key}.json'
    with open(out_file, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, ensure_ascii=False, indent=2)
    return {
        'file_key': file_key,
        'out_path': str(out_file),
        'rows': json_data,
        'header_str': header_str,
        'title': product_title,
    }


if __name__ == '__main__':
    import sys
    import glob

    # 默认路径
    excel_dir = Path('data/甲方excel')
    output_dir = 'data/input/parsed'

    # 找到甲方 Excel 文件
    excel_files = list(excel_dir.glob('*.xlsx'))
    if not excel_files:
        print(f'❌ 未在 {excel_dir} 找到 .xlsx 文件')
        sys.exit(1)

    excel_path = str(excel_files[0])
    print(f'📄 Excel 文件：{excel_path}')

    # 清空旧 JSON
    old_files = list(Path(output_dir).glob('*.json'))
    if old_files:
        for f in old_files:
            f.unlink()
        print(f'🗑️  已清空 {len(old_files)} 个旧 JSON 文件')
    else:
        print('ℹ️  output_dir 为空，无需清除')

    # 执行解析
    parse_excel(excel_path, output_dir)
