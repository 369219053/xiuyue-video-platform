"""
台账管理模块
负责创建、读取、更新 Excel 台账（Sheet1总览 / Sheet2成品清单 / Sheet3失败列表）
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from pathlib import Path
from datetime import datetime


# ─── 列定义 ───────────────────────────────────────────────────────────────────
COLUMNS = [
    'id', '组名', '日期', '人物a', '人物b', '场景', '道具',
    '台词', '字幕台词', '图片提示词', '视频提示词',
    '图片URL', '原始视频URL', '处理后视频URL', '最终视频URL',
    '当前阶段', '状态', '失败原因', '本地文件名', '废片标记'
]

STATUS_PENDING    = '⬜待处理'
STATUS_PROCESSING = '⏳处理中'
STATUS_SUCCESS    = '✅成功'
STATUS_FAILED     = '❌失败'


def _col_idx(name: str) -> int:
    """列名 → 1-based 列号"""
    return COLUMNS.index(name) + 1


def _make_header_style(ws, row: int):
    """给标题行加样式"""
    fill = PatternFill(fill_type='solid', fgColor='4472C4')
    font = Font(color='FFFFFF', bold=True)
    for col in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')


def create_ledger(output_path: str, entries: list[dict]) -> None:
    """
    初始化台账 Excel
    entries 格式：每条解析好的数据行字典，字段与 COLUMNS 对应
    """
    wb = openpyxl.Workbook()

    # ── Sheet1：总览 ──────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = '总览'
    ws1.append(COLUMNS)
    _make_header_style(ws1, 1)
    ws1.freeze_panes = 'A2'

    for entry in entries:
        row = [entry.get(col, '') for col in COLUMNS]
        ws1.append(row)

    # ── Sheet2：成品清单（给甲方）────────────────────────────────────────────
    ws2 = wb.create_sheet('成品清单')
    s2_cols = ['文件名', '视频URL', '图片URL', '组名', '日期', '台词摘要', '废片标记']
    ws2.append(s2_cols)
    _make_header_style(ws2, 1)

    # ── Sheet3：失败列表 ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet('失败列表')
    s3_cols = ['id', '失败阶段', '失败原因', '组名', '台词摘要']
    ws3.append(s3_cols)
    _make_header_style(ws3, 1)

    # 设置列宽
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"  📋 台账已创建：{output_path}（{len(entries)} 条）")


def update_entry(ledger_path: str, entry_id: str, fields: dict) -> None:
    """
    更新台账 Sheet1 中指定 id 的条目
    fields: {'图片URL': '...', '状态': '✅成功', ...}
    """
    wb = openpyxl.load_workbook(ledger_path)
    ws = wb['总览']

    id_col = _col_idx('id')
    for row in ws.iter_rows(min_row=2):
        if row[id_col - 1].value == entry_id:
            for field, value in fields.items():
                if field in COLUMNS:
                    row[_col_idx(field) - 1].value = value
            break

    wb.save(ledger_path)


def batch_update(ledger_path: str, updates: list[dict]) -> None:
    """
    批量更新台账（避免频繁读写文件）
    updates: [{'id': 'CXW_001', '图片URL': '...', '状态': '✅成功'}, ...]
    """
    wb = openpyxl.load_workbook(ledger_path)
    ws = wb['总览']
    id_col = _col_idx('id')

    # 建立 id → row 映射
    id_row_map = {}
    for row in ws.iter_rows(min_row=2):
        cell_id = row[id_col - 1].value
        if cell_id:
            id_row_map[cell_id] = row

    for update in updates:
        entry_id = update.get('id')
        if entry_id not in id_row_map:
            continue
        row = id_row_map[entry_id]
        for field, value in update.items():
            if field == 'id':
                continue
            if field in COLUMNS:
                row[_col_idx(field) - 1].value = value

    wb.save(ledger_path)
    print(f"  📋 台账已更新 {len(updates)} 条")


def finalize_ledger(ledger_path: str) -> None:
    """
    整理台账：填充 Sheet2（成品清单）和 Sheet3（失败列表）
    在所有工作流全部跑完后调用
    """
    wb = openpyxl.load_workbook(ledger_path)
    ws1 = wb['总览']
    ws2 = wb['成品清单']
    ws3 = wb['失败列表']

    for row in ws1.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        entry = dict(zip(COLUMNS, row))
        lyric_preview = str(entry.get('台词', ''))[:20]

        if entry.get('状态') == STATUS_SUCCESS:
            ws2.append([
                entry.get('本地文件名', ''),
                entry.get('最终视频URL', ''),
                entry.get('图片URL', ''),
                entry.get('组名', ''),
                entry.get('日期', ''),
                lyric_preview,
                ''  # 废片标记（甲方填写）
            ])
        elif entry.get('状态') == STATUS_FAILED:
            ws3.append([
                entry.get('id', ''),
                entry.get('当前阶段', ''),
                entry.get('失败原因', ''),
                entry.get('组名', ''),
                lyric_preview
            ])

    wb.save(ledger_path)
    print(f"  📋 台账已整理完成（成品清单 + 失败列表）")


def read_retry_ids(ledger_path: str) -> list[str]:
    """读取 Sheet2 中甲方标记了废片（废片标记='Y'）的条目 id"""
    wb = openpyxl.load_workbook(ledger_path)
    ws2 = wb['成品清单']
    s2_cols = ['文件名', '视频URL', '图片URL', '组名', '日期', '台词摘要', '废片标记']

    retry_ids = []
    for row in ws2.iter_rows(min_row=2, values_only=True):
        entry = dict(zip(s2_cols, row))
        if str(entry.get('废片标记', '')).strip().upper() == 'Y':
            # 文件名格式：CXW_001_20250515.mp4 → id = CXW_001_20250515
            filename = str(entry.get('文件名', ''))
            entry_id = filename.replace('.mp4', '')
            if entry_id:
                retry_ids.append(entry_id)

    print(f"  🔁 检测到废片 {len(retry_ids)} 条：{retry_ids}")
    return retry_ids
